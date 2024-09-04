import pytesseract
import cv2
import pyautogui
import time
import pandas as pd
import openpyxl
import win32gui
import win32com.client
import win32con


def pin_code(TesseractOCR: str):
    # pyautogui.press('f11')
    time.sleep(0.3)

    # Если tesseract не в PATH, укажите путь к исполняемому файлу
    pytesseract.pytesseract.tesseract_cmd = rf"{TesseractOCR}"

    # Сделать скриншот
    screenshot = pyautogui.screenshot(imageFilename='pc_vision/pin_code.jpg')

    # Открыть изображение с помощью OpenCV
    img = cv2.imread('pc_vision/pin_code.jpg')

    # Преобразовать изображение в оттенки серого
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.bilateralFilter(gray, 11, 17, 17)
    gray = cv2.Canny(gray, 30, 200)

    custom_config = r'--oem 3 --psm 6'

    # Распознать текст с координатами
    data = pytesseract.image_to_data(gray, lang='rus+eng', 
        output_type=pytesseract.Output.DICT, config=custom_config)

    # Найти индекс слова "руслан"
    ruslans_indices = [i for i, word in enumerate(data['text']) if 'руслан' in word.lower()]

    if ruslans_indices:
        ruslan_index = ruslans_indices[0]
        
        # Найти и кликнуть на "1" один раз
        for i in range(ruslan_index + 1, len(data['text'])):
            if '1' in data['text'][i]:
                x = data['left'][i]
                y = data['top'][i]
                w = data['width'][i]
                h = data['height'][i]
                
                center_x = x + w // 2
                center_y = y + h // 2
                
                pyautogui.click(center_x, center_y)
                # print(f"Clicked on 1 at ({center_x}, {center_y})")
                time.sleep(0.5)
                break

        # Найти "9" и кликнуть на неё три раза
        for i in range(ruslan_index + 1, len(data['text'])):
            if '9' in data['text'][i]:
                x = data['left'][i]
                y = data['top'][i]
                w = data['width'][i]
                h = data['height'][i]
                
                center_x = x + w // 2
                center_y = y + h // 2
                
                for _ in range(3):
                    pyautogui.click(center_x, center_y)
                    # print(f"Clicked on 9 at ({center_x}, {center_y})")
                    time.sleep(0.5)
                break

        return time.sleep(4)
    else:
        return None

# чтение excel файла и выдать нам list

def process_excel(file_path: str):
    column1_index = 0
    column2_index = 1

    # Читаем Excel файл без использования заголовков и без использования первого столбца как индекса
    df = pd.read_excel(file_path, header=None, index_col=None)
    
    # Создаем пустой список для результатов
    result_list = []
    
    # Проходим по строкам DataFrame
    for index, row in df.iterrows():
        value1 = row[column1_index]
        
        # Проверяем, есть ли второй столбец
        if df.shape[1] > column2_index:
            # Если есть второй столбец и значение в нем пустое, добавляем значение из первой колонки в список
            if pd.isna(row[column2_index]):
                result_list.append(value1)
        else:
            # Если второго столбца нет, просто добавляем значение из первой колонки
            result_list.append(value1)
    
    return result_list


# добавление новых значений в excel файл

def add_value_to_excel(file_path, search_value, value_to_add):
    search_value = int(search_value)
    column1_index=0
    column2_index=1

    # Открываем Excel файл
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    value_found = False
    
    # Проходим по строкам
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=column1_index + 1).value
        # Если нашли искомое значение в первой колонке
        if cell_value == search_value:
            # Добавляем новое значение во вторую колонку
            sheet.cell(row=row, column=column2_index + 1, value=value_to_add)
            # print(f"Added '{value_to_add}' next to '{search_value}' in row {row}")
            value_found = True
            break  # Прекращаем поиск после первого совпадения
    
    # if not value_found:
    #     print(f"Value '{search_value}' not found in the first column.")
    
    # Сохраняем изменения
    workbook.save(file_path)
    # print("Changes saved to the file.")


# активация приложения
def find_window_wildcard(wildcard):
    def callback(hwnd, hwnds):
        if win32gui.IsWindowVisible(hwnd) and win32gui.IsWindowEnabled(hwnd):
            title = win32gui.GetWindowText(hwnd)
            if wildcard.lower() in title.lower():
                hwnds.append(hwnd)
        return True
    hwnds = []
    win32gui.EnumWindows(callback, hwnds)
    return hwnds

def activate_window(window_title):
    try:
        hwnds = find_window_wildcard(window_title)
        if hwnds:
            shell = win32com.client.Dispatch("WScript.Shell")
            shell.SendKeys('%')
            win32gui.SetForegroundWindow(hwnds[0])
            return True

        return False

    except Exception as e:
        print(f"Не удалось активировать окно: {e}")
        return False
